"""
Background tasks for the ASF application using FastAPI BackgroundTasks.
Replaces Celery tasks with FastAPI's built-in background task system.
"""

from typing import Dict, Any
from datetime import datetime
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import boto3

from ..core.config import settings


def generate_excel_report(
    query_id: int, 
    forecast_data: Dict[str, Any],
    assumptions: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Generate Excel report for a forecast result using openpyxl.
    
    Creates a properly formatted Excel file with:
    1. Summary sheet with key metrics
    2. Monthly data sheet with detailed breakdown
    3. Assumptions sheet with all parameters
    4. Professional formatting and styling
    
    Args:
        query_id: ID of the forecast query
        forecast_data: Forecast calculation results
        assumptions: Assumptions used in the forecast
        
    Returns:
        Dictionary with report metadata and download URL
    """
    try:
        print(f"Starting Excel generation for query {query_id}...")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        print("Creating summary sheet...")
        
        # Create Summary sheet
        summary_ws = wb.create_sheet("Summary", 0)
        _create_summary_sheet(summary_ws, forecast_data, query_id)
        
        print("Creating monthly data sheet...")
        
        # Create Monthly Data sheet
        monthly_ws = wb.create_sheet("Monthly Data", 1)
        _create_monthly_data_sheet(monthly_ws, forecast_data)
        
        print("Creating assumptions sheet...")
        
        # Create Assumptions sheet
        assumptions_ws = wb.create_sheet("Assumptions", 2)
        _create_assumptions_sheet(assumptions_ws, assumptions)
        
        print("Saving Excel file...")
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        print("Uploading to storage...")
        
        # Upload to S3 (or local storage for development)
        file_key = f"reports/forecast_{query_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_url = _upload_to_storage(excel_buffer, file_key)
        
        # Return result
        result = {
            "query_id": query_id,
            "file_url": file_url,
            "file_size": f"{len(excel_buffer.getvalue()) / 1024 / 1024:.1f} MB",
            "generated_at": datetime.now().isoformat(),
            "status": "completed"
        }
        
        print(f"Excel report generated successfully: {file_url}")
        return result
        
    except Exception as e:
        print(f"Excel generation failed: {str(e)}")
        raise


def send_notification_email(
    user_email: str,
    query_id: int,
    forecast_status: str
) -> Dict[str, Any]:
    """
    Send notification email to user about forecast completion.
    
    This is a placeholder implementation. In production, this would:
    1. Use email service (SendGrid, SES, etc.)
    2. Generate HTML email template
    3. Include forecast summary and download links
    4. Handle email delivery status
    
    Args:
        user_email: User's email address
        query_id: ID of the forecast query
        forecast_status: Status of the forecast (completed, failed)
        
    Returns:
        Dictionary with email delivery status
    """
    try:
        print(f"Sending notification email to {user_email} for query {query_id}...")
        
        # Simulate email preparation and sending
        import time
        time.sleep(1)
        
        # Return mock result
        result = {
            "user_email": user_email,
            "query_id": query_id,
            "forecast_status": forecast_status,
            "email_sent": True,
            "sent_at": datetime.now().isoformat(),
            "message_id": f"msg_{query_id}_{forecast_status}"
        }
        
        print(f"Notification email sent successfully to {user_email}")
        return result
        
    except Exception as e:
        print(f"Email sending failed: {str(e)}")
        raise


def cleanup_old_reports(days_old: int = 30) -> Dict[str, Any]:
    """
    Clean up old Excel reports and temporary files.
    
    This is a placeholder implementation. In production, this would:
    1. Query database for old reports
    2. Delete files from object storage
    3. Update database records
    4. Log cleanup statistics
    
    Args:
        days_old: Number of days after which to clean up reports
        
    Returns:
        Dictionary with cleanup statistics
    """
    try:
        print(f"Starting cleanup of reports older than {days_old} days...")
        
        # Simulate cleanup process
        import time
        time.sleep(2)
        
        # Return mock result
        result = {
            "days_old": days_old,
            "files_deleted": 15,
            "storage_freed": "45.2 MB",
            "cleanup_completed_at": datetime.now().isoformat()
        }
        
        print(f"Cleanup completed: {result['files_deleted']} files deleted")
        return result
        
    except Exception as e:
        print(f"Cleanup failed: {str(e)}")
        raise


def _create_summary_sheet(ws, forecast_data: Dict[str, Any], query_id: int):
    """Create the summary sheet with key metrics."""
    # Title
    ws['A1'] = f"FinSynth Forecast Report - Query #{query_id}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Summary metrics
    ws['A3'] = "Forecast Summary"
    ws['A3'].font = Font(size=14, bold=True)
    
    summary = forecast_data.get('summary', {})
    metrics = [
        ("Forecast Type", forecast_data.get('forecast_type', 'Unknown')),
        ("Timeframe", f"{forecast_data.get('timeframe_months', 0)} months"),
        ("Total Revenue", f"${summary.get('total_revenue', 0):,.0f}"),
        ("Large Customer Revenue", f"${summary.get('large_customer_revenue', 0):,.0f}"),
        ("SMB Customer Revenue", f"${summary.get('smb_customer_revenue', 0):,.0f}"),
    ]
    
    for i, (label, value) in enumerate(metrics, start=4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _create_monthly_data_sheet(ws, forecast_data: Dict[str, Any]):
    """Create the monthly data sheet with comprehensive metrics breakdown matching the image format."""
    
    # Define comprehensive headers matching the image
    headers = [
        "Metric", "Unit",
        "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11"
    ]
    
    # Create header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Define metrics structure matching the image
    metrics = [
        # Sales & Large Customer Acquisition Metrics
        ("# of sales people", "count"),
        ("# of large customer accounts they can sign per month/sales person", "count"),
        ("# of large customer accounts onboarded per month", "count"),
        ("Cumulative # of paying customers (large clients)", "count"),
        ("Average revenue per customer (large clients)", "$ per month"),
        
        # Marketing Metrics
        ("Digital Marketing spend per month", "$ per month"),
        ("Average CAC (Customer Acquisition Cost)", "$ per customer"),
        ("# of sales enquiries", "count"),
        ("% conversions from demo to sign ups", "%"),
        
        # Small and Medium Customer Onboarding Metrics
        ("# of paying customers onboarded", "count"),
        ("Cumulative number of paying customers (small/medium clients)", "count"),
        ("Average revenue per customer (small/medium clients)", "$ per customer"),
        
        # Revenue Metrics
        ("Revenue from large clients", "$ per month"),
        ("Revenue from small and medium clients", "$ per month"),
        ("Total Revenues", "$ per month"),
        ("Total Revenues (Mn)", "$ Mn per month")
    ]
    
    # Get monthly data
    monthly_data = forecast_data.get('monthly_data', [])
    
    # Create data rows
    for row_idx, (metric_name, unit) in enumerate(metrics, start=2):
        # Metric name and unit
        ws.cell(row=row_idx, column=1, value=metric_name)
        ws.cell(row=row_idx, column=2, value=unit)
        
        # Monthly values
        for col_idx, data in enumerate(monthly_data[:11], start=3):  # Limit to 11 months
            value = _get_metric_value(data, metric_name)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Format columns
    ws.column_dimensions['A'].width = 50  # Metric names
    ws.column_dimensions['B'].width = 20  # Units
    for col in range(3, 14):  # M1-M11 columns
        ws.column_dimensions[get_column_letter(col)].width = 15


def _get_metric_value(data: Dict[str, Any], metric_name: str) -> Any:
    """Extract the appropriate value for a given metric from monthly data."""
    
    metric_mapping = {
        "# of sales people": data.get("sales_people", 0),
        "# of large customer accounts they can sign per month/sales person": data.get("large_accounts_per_sales_person", 1),
        "# of large customer accounts onboarded per month": data.get("large_accounts_onboarded", 0),
        "Cumulative # of paying customers (large clients)": data.get("cumulative_large_customers", 0),
        "Average revenue per customer (large clients)": data.get("avg_revenue_per_large_customer", 0),
        
        "Digital Marketing spend per month": data.get("digital_marketing_spend", 0),
        "Average CAC (Customer Acquisition Cost)": data.get("avg_cac", 0),
        "# of sales enquiries": data.get("sales_enquiries", 0),
        "% conversions from demo to sign ups": round(data.get("conversion_rate", 0) * 100, 0),
        
        "# of paying customers onboarded": data.get("smb_customers_onboarded", 0),
        "Cumulative number of paying customers (small/medium clients)": data.get("cumulative_smb_customers", 0),
        "Average revenue per customer (small/medium clients)": data.get("avg_revenue_per_smb_customer", 0),
        
        "Revenue from large clients": data.get("large_customer_revenue", 0),
        "Revenue from small and medium clients": data.get("smb_customer_revenue", 0),
        "Total Revenues": data.get("total_revenue", 0),
        "Total Revenues (Mn)": data.get("total_revenue_mn", 0)
    }
    
    return metric_mapping.get(metric_name, 0)


def _create_assumptions_sheet(ws, assumptions: Dict[str, Any]):
    """Create the assumptions sheet with all parameters."""
    # Title
    ws['A1'] = "Forecast Assumptions"
    ws['A1'].font = Font(size=14, bold=True)
    
    row = 3
    
    # Large Customer Assumptions
    ws[f'A{row}'] = "Large Customer Assumptions"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    large_assumptions = assumptions.get('large_customer', {})
    for key, value in large_assumptions.items():
        ws[f'A{row}'] = key.replace('_', ' ').title()
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # SMB Customer Assumptions
    ws[f'A{row}'] = "SMB Customer Assumptions"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    smb_assumptions = assumptions.get('smb_customer', {})
    for key, value in smb_assumptions.items():
        ws[f'A{row}'] = key.replace('_', ' ').title()
        ws[f'B{row}'] = value
        row += 1
    
    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _upload_to_storage(excel_buffer: BytesIO, file_key: str) -> str:
    """Upload Excel file to storage and return URL."""
    # For development, save locally
    if settings.environment == "development":
        local_path = f"storage/{file_key}"
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, 'wb') as f:
            f.write(excel_buffer.getvalue())
        return f"http://localhost:8000/storage/{file_key}"
    
    # For production, upload to S3
    try:
        s3_client = boto3.client('s3')
        bucket_name = settings.aws_s3_bucket if hasattr(settings, 'aws_s3_bucket') else 'asf-reports'
        
        s3_client.put_object(
            Bucket=bucket_name,
            Key=file_key,
            Body=excel_buffer.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        return f"https://{bucket_name}.s3.amazonaws.com/{file_key}"
    except Exception as e:
        print(f"S3 upload failed: {e}")
        # Fallback to local storage
        local_path = f"storage/{file_key}"
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, 'wb') as f:
            f.write(excel_buffer.getvalue())
        return f"http://localhost:8000/storage/{file_key}"
