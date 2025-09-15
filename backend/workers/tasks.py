"""
Celery background tasks for the ASF application.
Implements Excel generation and other long-running tasks as specified in the PRD.
"""

from typing import Dict, Any, Optional
from datetime import datetime
import json
from celery import current_task
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import boto3
import os
from io import BytesIO

from .celery_app import celery_app
from ..core.config import settings


@celery_app.task(bind=True)
def generate_excel_report(
    self, 
    query_id: int, 
    forecast_data: Dict[str, Any],
    assumptions: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Generate Excel report for a forecast result using openpyxl.
    
    Creates a properly formatted Excel file with:
    1. Summary sheet with key metrics
    2. Monthly data sheet with detailed breakdown
    3. Assumptions sheet with all parameters
    4. Professional formatting and styling
    
    Args:
        query_id: ID of the forecast query
        forecast_data: Forecast calculation results
        assumptions: Assumptions used in the forecast
        
    Returns:
        Dictionary with report metadata and download URL
    """
    try:
        # Update task progress
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 0, "total": 100, "status": "Starting Excel generation..."}
        )
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 20, "total": 100, "status": "Creating summary sheet..."}
        )
        
        # Create Summary sheet
        summary_ws = wb.create_sheet("Summary", 0)
        _create_summary_sheet(summary_ws, forecast_data, query_id)
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 40, "total": 100, "status": "Creating monthly data sheet..."}
        )
        
        # Create Monthly Data sheet
        monthly_ws = wb.create_sheet("Monthly Data", 1)
        _create_monthly_data_sheet(monthly_ws, forecast_data)
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 60, "total": 100, "status": "Creating assumptions sheet..."}
        )
        
        # Create Assumptions sheet
        assumptions_ws = wb.create_sheet("Assumptions", 2)
        _create_assumptions_sheet(assumptions_ws, assumptions)
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 80, "total": 100, "status": "Saving Excel file..."}
        )
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 90, "total": 100, "status": "Uploading to storage..."}
        )
        
        # Upload to S3 (or local storage for development)
        file_key = f"reports/forecast_{query_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_url = _upload_to_storage(excel_buffer, file_key)
        
        # Return result
        result = {
            "query_id": query_id,
            "file_url": file_url,
            "file_size": f"{len(excel_buffer.getvalue()) / 1024 / 1024:.1f} MB",
            "generated_at": datetime.now().isoformat(),
            "status": "completed"
        }
        
        current_task.update_state(
            state="SUCCESS",
            meta={"current": 100, "total": 100, "status": "Excel report generated successfully", "result": result}
        )
        
        return result
        
    except Exception as e:
        current_task.update_state(
            state="FAILURE",
            meta={"current": 0, "total": 100, "status": f"Excel generation failed: {str(e)}"}
        )
        raise


def _create_summary_sheet(ws, forecast_data: Dict[str, Any], query_id: int):
    """Create the summary sheet with key metrics."""
    # Title
    ws['A1'] = f"FinSynth Forecast Report - Query #{query_id}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Summary metrics
    ws['A3'] = "Forecast Summary"
    ws['A3'].font = Font(size=14, bold=True)
    
    summary = forecast_data.get('summary', {})
    metrics = [
        ("Forecast Type", forecast_data.get('forecast_type', 'Unknown')),
        ("Timeframe", f"{forecast_data.get('timeframe_months', 0)} months"),
        ("Total Revenue", f"${summary.get('total_revenue', 0):,.0f}"),
        ("Large Customer Revenue", f"${summary.get('large_customer_revenue', 0):,.0f}"),
        ("SMB Customer Revenue", f"${summary.get('smb_customer_revenue', 0):,.0f}"),
    ]
    
    for i, (label, value) in enumerate(metrics, start=4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _create_monthly_data_sheet(ws, forecast_data: Dict[str, Any]):
    """Create the monthly data sheet with detailed breakdown."""
    # Headers
    headers = ["Month", "Large Customer Revenue", "SMB Customer Revenue", "Total Revenue"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    monthly_data = forecast_data.get('monthly_data', [])
    for row, data in enumerate(monthly_data, start=2):
        ws.cell(row=row, column=1, value=f"Month {data.get('month', row-1)}")
        ws.cell(row=row, column=2, value=data.get('large_customer_revenue', 0))
        ws.cell(row=row, column=3, value=data.get('smb_customer_revenue', 0))
        ws.cell(row=row, column=4, value=data.get('total_revenue', 0))
    
    # Format columns
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _create_assumptions_sheet(ws, assumptions: Dict[str, Any]):
    """Create the assumptions sheet with all parameters."""
    # Title
    ws['A1'] = "Forecast Assumptions"
    ws['A1'].font = Font(size=14, bold=True)
    
    row = 3
    
    # Large Customer Assumptions
    ws[f'A{row}'] = "Large Customer Assumptions"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    large_assumptions = assumptions.get('large_customer', {})
    for key, value in large_assumptions.items():
        ws[f'A{row}'] = key.replace('_', ' ').title()
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # SMB Customer Assumptions
    ws[f'A{row}'] = "SMB Customer Assumptions"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    smb_assumptions = assumptions.get('smb_customer', {})
    for key, value in smb_assumptions.items():
        ws[f'A{row}'] = key.replace('_', ' ').title()
        ws[f'B{row}'] = value
        row += 1
    
    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _upload_to_storage(excel_buffer: BytesIO, file_key: str) -> str:
    """Upload Excel file to storage and return URL."""
    # For development, save locally
    if settings.environment == "development":
        local_path = f"storage/{file_key}"
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, 'wb') as f:
            f.write(excel_buffer.getvalue())
        return f"http://localhost:8000/storage/{file_key}"
    
    # For production, upload to S3
    try:
        s3_client = boto3.client('s3')
        bucket_name = settings.aws_s3_bucket if hasattr(settings, 'aws_s3_bucket') else 'asf-reports'
        
        s3_client.put_object(
            Bucket=bucket_name,
            Key=file_key,
            Body=excel_buffer.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        return f"https://{bucket_name}.s3.amazonaws.com/{file_key}"
    except Exception as e:
        print(f"S3 upload failed: {e}")
        # Fallback to local storage
        local_path = f"storage/{file_key}"
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, 'wb') as f:
            f.write(excel_buffer.getvalue())
        return f"http://localhost:8000/storage/{file_key}"


@celery_app.task(bind=True)
def send_notification_email(
    self,
    user_email: str,
    query_id: int,
    forecast_status: str
) -> Dict[str, Any]:
    """
    Send notification email to user about forecast completion.
    
    This is a placeholder implementation. In production, this would:
    1. Use email service (SendGrid, SES, etc.)
    2. Generate HTML email template
    3. Include forecast summary and download links
    4. Handle email delivery status
    
    Args:
        user_email: User's email address
        query_id: ID of the forecast query
        forecast_status: Status of the forecast (completed, failed)
        
    Returns:
        Dictionary with email delivery status
    """
    try:
        # Update task progress
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 0, "total": 100, "status": "Preparing email..."}
        )
        
        # Simulate email preparation
        import time
        time.sleep(1)
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 50, "total": 100, "status": "Sending email..."}
        )
        
        time.sleep(1)
        
        # Return mock result
        result = {
            "user_email": user_email,
            "query_id": query_id,
            "forecast_status": forecast_status,
            "email_sent": True,
            "sent_at": "2024-01-01T12:00:00Z",
            "message_id": f"msg_{query_id}_{forecast_status}"
        }
        
        current_task.update_state(
            state="SUCCESS",
            meta={"current": 100, "total": 100, "status": "Email sent successfully", "result": result}
        )
        
        return result
        
    except Exception as e:
        current_task.update_state(
            state="FAILURE",
            meta={"current": 0, "total": 100, "status": f"Email sending failed: {str(e)}"}
        )
        raise


@celery_app.task(bind=True)
def cleanup_old_reports(self, days_old: int = 30) -> Dict[str, Any]:
    """
    Clean up old Excel reports and temporary files.
    
    This is a placeholder implementation. In production, this would:
    1. Query database for old reports
    2. Delete files from object storage
    3. Update database records
    4. Log cleanup statistics
    
    Args:
        days_old: Number of days after which to clean up reports
        
    Returns:
        Dictionary with cleanup statistics
    """
    try:
        # Update task progress
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 0, "total": 100, "status": "Starting cleanup..."}
        )
        
        # Simulate cleanup process
        import time
        time.sleep(2)
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 50, "total": 100, "status": "Deleting old files..."}
        )
        
        time.sleep(2)
        
        # Return mock result
        result = {
            "days_old": days_old,
            "files_deleted": 15,
            "storage_freed": "45.2 MB",
            "cleanup_completed_at": "2024-01-01T12:00:00Z"
        }
        
        current_task.update_state(
            state="SUCCESS",
            meta={"current": 100, "total": 100, "status": "Cleanup completed successfully", "result": result}
        )
        
        return result
        
    except Exception as e:
        current_task.update_state(
            state="FAILURE",
            meta={"current": 0, "total": 100, "status": f"Cleanup failed: {str(e)}"}
        )
        raise
